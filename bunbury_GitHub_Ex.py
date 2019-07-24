'''
This python script reads an Excel workbook and generates an XML file
The XML file is compatible with the Tilia paleoecology software program
Tilia program enables review of data before entry to Neotoma database

Full documentation available here:
https://github.com/NeotomaDB/xlsx2Tilia

This program was supported by NSF 1550890

Comments can be emailed to author: Andrew Anderson at acandrsn@illinois.edu
'''


import pandas as pd
import xml.etree.cElementTree as et
from openpyxl import load_workbook

# get the number of rows in the dataset
file_name_string = 'data/bunbury_input.xlsx'
name_of_tab = "mainTable"
dfMainTable = pd.read_excel(file_name_string, sheet_name=name_of_tab, index_col=0)
maxLoop = len(dfMainTable.index)
print(maxLoop)

def createTilia(bigLoopCounter):
    # load workbook for use with openpyxl
    wb = load_workbook(filename='data/bunbury_input.xlsx', read_only=True, data_only=True)

    # use with pandas pd.read_excel
    file_name_string = 'data/bunbury_input.xlsx'

    # define function for later
    # read a range with openpyxl, return a dataframe
    def returnRangeValues(rangeStart,rangeStop):

        data_rows = []
        for row in ws[rangeStart:rangeStop]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)

        # Transform into dataframe
        df = pd.DataFrame(data_rows)
        # get the number of rows in the dataframe
        return df

    def isanumber(a):
        # note: isdecimal(), isnumeric(), isdigit() don't work for all numbers w/ unicode problems

        try:
            float(repr(a))
            bool_a = True
        except:
            bool_a = False

        return bool_a

    # pandas assumes the first row are headers = keys
    # set the index column to the first column
    # ===================================================
    #                 get sheet = contacts
    #                 dataframe = myData
    #                 updating = no
    # ===================================================
    myData = pd.read_excel(file_name_string, sheet_name='contacts', index_col=0)


    # ===================================================
    #             Start XML file generation
    # ===================================================

    root = et.Element("TiliaFile")  # create a root element
    sub = et.SubElement(root, "Version")  # create sub under root (when added ID=1 there is an output error)

    et.SubElement(sub, "Application").text = "Tilia"
    et.SubElement(sub, "MajorVersion").text = "2"
    et.SubElement(sub, "MinorVersion").text = "0"
    et.SubElement(sub, "Release").text = "41"

    # ===================================================
    #             XML Contacts Generation
    # ===================================================

    rowCount = len(myData.index)
    #print("The number of records is:", rowCount)

    contactCount = et.SubElement(root,"Contacts")

    # i = column
    # j = row
    # loop through rows

    for j in range(0, len(myData.index)):
        # create a new contact XML group
        contact = et.SubElement(contactCount, "Contact", ID=str(j + 1))

        # loop through columns
        for i in range(0,len(myData.columns)):
            # catch potential blanks = "nan"
            if pd.isnull(myData.iloc[j,i]) == False:
                if i < 2:
                    # must be converted to int before string to prevent e.g. 123.0 for research IDs
                    et.SubElement(contact, myData.columns[i]).text = str(int(myData.iloc[j, i]))
                    #print(str(int(myData.iloc[j, i])), "where i =",i,"and j=",j)
                elif i >= 2 and i < 13:
                    # for j < 13
                    et.SubElement(contact, myData.columns[i]).text = str(myData.iloc[j, i])
                elif i == 13:
                    addressgroup = et.SubElement(contact, "Address")
                    et.SubElement(addressgroup, "AddressLine").text = str(myData.iloc[j,i])
                elif i > 13:
                    et.SubElement(addressgroup, "AddressLine").text = str(myData.iloc[j,i])



    # ===================================================
    #                 get sheet = xlsFormat
    #                 dataframe = tableProps
    #                 updating = no
    # ===================================================
    tableProps = pd.read_excel(file_name_string, sheet_name='xlsFormat', index_col=None)

    # ===================================================
    #             XML Tilia general attributes Generation
    # ===================================================

    #order is [row, column] with index starting at 0
    xlsBook = et.SubElement(root, "SpreadSheetBook")
    xlsOptions = et.SubElement(xlsBook, "SpreadSheetOptions")
    et.SubElement(xlsOptions, "HeaderRow").text = "0"
    et.SubElement(xlsOptions, "FontName").text = str(tableProps.iloc[0,1])
    et.SubElement(xlsOptions, "FontSize").text = str(tableProps.iloc[0,2])
    et.SubElement(xlsOptions, "DefaultColWidth").text = str(tableProps.iloc[0,3])
    et.SubElement(xlsOptions, "DefaultRowHeight").text = str(tableProps.iloc[0,4])
    et.SubElement(xlsOptions, "PercentDecimalPlaces").text = str(tableProps.iloc[0,5])
    et.SubElement(xlsOptions, "CheckDupCodes").text = str(tableProps.iloc[0,6])
    et.SubElement(xlsOptions, "CaseSensitiveCodes").text = str(tableProps.iloc[0,7])
    et.SubElement(xlsOptions, "CodesVisible").text = str(tableProps.iloc[0,8])
    et.SubElement(xlsOptions, "ElementsVisible").text = str(tableProps.iloc[0,9])
    et.SubElement(xlsOptions, "UnitsVisible").text = str(tableProps.iloc[0,10])
    et.SubElement(xlsOptions, "ContextsVisible").text = str(tableProps.iloc[0,11])
    et.SubElement(xlsOptions, "TaphonomyVisible").text = str(tableProps.iloc[0,12])
    et.SubElement(xlsOptions, "GroupsVisible").text = str(tableProps.iloc[0,13])
    # et.SubElement(xlsOptions, "ElementCategory").text = str(tableProps.iloc[0,14])  # not used in latest Tilia version
    # et.SubElement(xlsOptions, "UnitsCategory").text = str(tableProps.iloc[0,15])
    # et.SubElement(xlsOptions, "ContextCategory").text = str(tableProps.iloc[0,16])
    # et.SubElement(xlsOptions, "TaphonomyCategory").text = str(tableProps.iloc[0,17])

    # ===================================================
    #            Start of 'spreadsheet' table
    # ===================================================

    # ======================PART1========================
    #                 get sheet = ostracode
    #                 dataframe = dfostra
    #                 updating = yes
    # ===================================================
    # contains species only not water quality data
    # bring in data table
    # excel table has 2 header rows so set skiprows = 1 (first row considered header by default)
    dfOstra = pd.read_excel(file_name_string, sheet_name='ostracode', header=None, index_col=None)

    # ===================================================
    #          get ostracode presence/absence
    # ===================================================

    # read range to determine presence/absence of ostracode species
    # specify worksheet; data_only=True returns values not cell formulas
    ws = wb['mainTable']

    # returnDataframe = returnRangeValues('AV2', 'CE2')
    # add 1 to bigLoopCounter, data starts on row 2
    # AV to KS is the new presence absence range
    returnDataframe = returnRangeValues(("AV" + str(bigLoopCounter+1)), ("KS" + str(bigLoopCounter+1)))

    # create a list to hold positions of 0 values indicating ostracode species not present
    zeroList = []

    # fill list with index positions for values equal to 0 from mainTable Excel sheet
    # 0 = absent, 1 = present
    for i in range(0, len(returnDataframe.columns)):
        if returnDataframe.iloc[0, i] == 0:
            # There are 3 header rows = +3
            zeroList.append(i+3)

    # print("The list is:", zeroList)

    #removes species which are not present according to the main table
    dfOstraUpdate = dfOstra.drop(dfOstra.index[zeroList])

    # print(dfOstraUpdate)


    # '''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    # check calc with this
    # print(returnDataframe.head())
    # print(dfOstraUpdate)
    #
    # print("The shape of the dataframe is", returnDataframe.shape)
    # # 1,36# print("The number of rows in the dataframe is", len(returnDataframe.index))
    # print("The number of columns in the dataframe is", len(returnDataframe.columns))
    #
    # print("The value of the first position is", returnDataframe.iloc[0,0])
    #

    # writer = pd.ExcelWriter('reducedSpecies.xlsx', engine='openpyxl')
    # dfOstraUpdate.to_excel(writer, index=False)
    # writer.save()

    # end of check
    # '''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

    # # =====================PART2=========================
    # #                 get sheet = wChem
    # #                 dataframe = dfWCHEM
    # #                 updating = yes
    # # ===================================================
    # bring in data table
    # excel table has 2 header rows so set skiprows = 1

    dfWCHEM = pd.read_excel(file_name_string, sheet_name='wChem', header=None, index_col=None)

    # ``````````````````````````` Updating code ``````````````````````````
    # read range to overwrite data in template
    # specify worksheet; data_only=True returns values not cell formulas
    ws = wb['mainTable']

    # set range values
    # add value = 1 to bigLoopCounter, data starts on row 2
    returnDataframe = returnRangeValues(("Z" + str(bigLoopCounter+1)), ("AU" + str(bigLoopCounter+1)))
    # returnDataframe = returnRangeValues('U2', 'AP2')
    # get the number of columns
    columnCount = len(returnDataframe.columns)
    # print("The number of columns in the return dataframe is,", columnCount)

    # index is offset by 1 because no header is included
    # first row is a header but must be used to create the XML
    # therefore cannot be treated as a header
    for n in range(0, 22):
        # change to column #9 (index #8, for Tilia 2.0 format)
        # sfWCHEM + 2 as there are 2 header rows in the original dataframe
        dfWCHEM.iloc[n+2,8]= returnDataframe.iloc[0,n]

    # print(dfWCHEM)

    # drop rows just from dfWCHEM
    dfWCHEM = dfWCHEM.drop(dfWCHEM.index[range(2, 8)])

    # `````````````````````````````````````````````````````````````````````

    # ===================================================
    #     Merge dfOstraUpdate and dfWCHEM dataframes
    # ===================================================
    # now the dfWCHEM table should be appended to the dfOstraUpdate table
    # dataframes must have the same column names
    frames = [dfOstraUpdate, dfWCHEM]
    # get names of columns from dfOstraUpdate
    columnNames = dfOstraUpdate.columns.values.tolist()
    # columnNames = [0,1,2....8]

    for df in frames:
        df.columns = columnNames

    # merged the dataframes, reset index
    dfData = pd.concat(frames).reset_index(drop=True)

    #reset values to get LDDsite#-OST and LDDsite#-WCH
    ws = wb['mainTable']

    #lddName1 = ws["c", bigLoopCounter+1].value
    lddName1 = ws["KT" + str(bigLoopCounter+1)].value   #gives Excel location e.g. CF2, CF3
    lddName2 = ws["KU" + str(bigLoopCounter+1)].value

    # print(dfData.head())

    # write value to row 1 and columns 7 and 8.  Write the site-OST and site-WCH values
    dfData.set_value(1, 7, lddName1)
    dfData.set_value(1, 8, lddName2)


    # test dataframe concat for water chemistry table
    writer = pd.ExcelWriter('tempOutput.xlsx')
    dfWCHEM.to_excel(writer, 'wChem', header=False, index=False)
    dfData.to_excel(writer, 'completeTable', header=False, index=False)
    writer.save()
    # end of test writing otu dataframe code

    # ===================================================
    #            XML 'spreadsheet' table creation
    # ===================================================

    # create data table in XML
    xlsName = et.SubElement(xlsBook, "SpreadSheet", name="Data", page="0")

    # skip 3 rows = (cellrowStart = 1) +
    cellRowStart = 1
    rowCountData = len(dfData.index)
    #print("The number of rows of data is:", rowCountData)

    columnCounter = 1
    # get the number of columns in the dataframe
    numColumns = len(dfData.columns)

    # we now are eliminating the temperature and precipitation data from Delorme
    # Neotoma provides this information database wide using specific data/model
    # dfDataNew = dfData.drop(dfData.index[range(6,12)])
    # # print(dfDataNew)
    # dfData = None
    # dfData = dfDataNew

    # updated row count
    rowCountData = len(dfData.index)

    for j in range(0, 7):
        xlsColumn = et.SubElement(xlsName, "Col", ID=str(j+1), Width="87")             #gives e.g <Col ID = 5 Width = "87">

        # loop over rows
        # if statement will skip over blank Excel cells which are read in as nan
        # skip first 2 rows (2 header rows, next row is #Samp.Analyst
        for i in range(2, rowCountData):
            if pd.isnull(dfData.iloc[i, j]) == False:
                if isanumber(dfData.iloc[i,j]) ==True:
                    xlsRow = et.SubElement(xlsColumn, "cell", row=str(i + cellRowStart))
                    xlsText = et.SubElement(xlsRow, "value").text = str(dfData.iloc[i, j])
                else:
                    xlsRow = et.SubElement(xlsColumn, "cell", row=str(i + cellRowStart))
                    # gives e.g. <cell row = "3">
                    xlsText = et.SubElement(xlsRow, "text").text = str(dfData.iloc[i, j])

                    # exceptions to the main table
                    # print("The value of i is:", i)
                    if (dfData.iloc[i,j]=="Bunbury, J."):
                        # creates the contactID tag
                        # overwrites previous "text"? if true?
                        et.SubElement(xlsRow, "contact", ID=str(2)).text = ""

                    if str(dfData.iloc[i,j])=="Neotoma Ostracode:visceral mass present":
                        # print("verceral mass column row", i)
                        xlsTaph = et.SubElement(xlsRow, "Taphonomy", System="Neotoma Ostracode", DatasetType="ostracode")
                        xlsType = et.SubElement(xlsTaph, "Type").text = "visceral mass present"
                        # below code manually closes out Taphonomy
                        # et.tostring(et.fromstring('<Taphonomy/>'), method='html')


            # this section below is for columns 7 and 8
            # the headers are not written out for columns 0 to 6
            # and therefore start at row 3
            # however for columns 7 and 8 1 header row is written and row = 2
    for j in range(7, numColumns):
        xlsColumn = et.SubElement(xlsName, "Col", ID=str(j + 1), Width="87")

        for i in range (1, rowCountData):
            if pd.isnull(dfData.iloc[i,j]) == False:
                if isanumber(dfData.iloc[i,j]) == True:
                    # function isanumber, custom b/c python defaults for number testing fail
                    # this catches that in version 2.0 for last columns those with numbers are now 'value' not 'text'
                    xlsRow = et.SubElement(xlsColumn, "cell", row=str(i + cellRowStart))
                    xlsText = et.SubElement(xlsRow, "value").text = str(dfData.iloc[i, j])
                else:
                    xlsRow = et.SubElement(xlsColumn, "cell", row=str(i+cellRowStart))
                    # gives e.g. <cell row = "3">
                    xlsText = et.SubElement(xlsRow, "text").text = str(dfData.iloc[i,j])

                    # exceptions to the main table
                    # print("The value of i is:", i)
                    if (dfData.iloc[i,j]=="Bunbury, J."):
                        # creates the contactID tag
                        # overwrites previous "text"? if true?
                        et.SubElement(xlsRow, "contact", ID=str(2)).text = ""

                    if str(dfData.iloc[i,j])=="Neotoma Ostracode:visceral mass present":
                        # print("verceral mass column row", i)
                        xlsTaph = et.SubElement(xlsRow, "Taphonomy", System="Neotoma Ostracode", DatasetType="ostracode")
                        xlsType = et.SubElement(xlsTaph, "Type").text = "visceral mass present"
                        # below code manually closes out Taphonomy
                        # et.tostring(et.fromstring('<Taphonomy/>'), method='html')

    # ===================================================
    #                 get sheet = site
    #                 dataframe = siteProps
    #                 updating = yes
    # ===================================================

    # note I put the data in columns here not rows.  So there is no header => header=None
    # load the template
    siteProps = pd.read_excel(file_name_string, sheet_name="site",header=None) # just used for the column names
    # using header = none puts the header in the first row = row 0
    rowCount_siteProps = len(siteProps.index)

    # ``````````````````````````` Updating code ``````````````````````````
    # read range to overwrite data in template
    # specify worksheet; data_only=True returns values not cell formulas
    ws = wb['mainTable']

    # set range values
    # add value = 1 to bigLoopCounter, data starts on row 2
    returnDataframe = returnRangeValues(("C" + str(bigLoopCounter+1)), ("Y" + str(bigLoopCounter+1)))
    #returnDataframe = returnRangeValues('C2', 'T2')
    # get the number of columns
    colNumber = len(returnDataframe.columns)

    for n in range (0, colNumber):
        siteProps.iloc[1,n]= returnDataframe.iloc[0,n]
    # `````````````````````````````````````````````````````````````````````

    # ===================================================
    #            XML 'site' table creation
    # ===================================================

    # write out the general site information
    # dataframe name = siteProps
    # subelement(name_upper_tag, name_tag_to_create).text = text between tags

    # ====== Site Section ==========================
    site = et.SubElement(root, "Site")
    # starts with SiteName
    # et.SubElement(site, "SiteName").text = str(siteProps.iloc[1, 0])
    for k in range(0, 10):
        if pd.isnull(siteProps.iloc[1, k]) == False:
            et.SubElement(site, str(siteProps.iloc[0,k])).text = str(siteProps.iloc[1,k])

    # ====== Collection Unit Section ================
    collectionUnit = et.SubElement(root, "CollectionUnit")
    # starts with 'Handle'
    for k in range(10, 14):
        if pd.isnull(siteProps.iloc[1, k]) == False:
            et.SubElement(collectionUnit, str(siteProps.iloc[0,k])).text = str(siteProps.iloc[1,k])

    collectCount = et.SubElement(collectionUnit, "Collectors")
    contactCount = et.SubElement(collectCount, "Contact", ID="2")

    # skip position 14, just a holder for collector
    for k in range(15, 19):
        if pd.isnull(siteProps.iloc[1, k]) == False:
            et.SubElement(collectionUnit, str(siteProps.iloc[0,k])).text = str(siteProps.iloc[1,k])

    # ===== Dataset Section ==========================
    datasetsVar = et.SubElement(root, "Datasets")
    datasetVar = et.SubElement(datasetsVar, "Dataset")

    #remember that the index says upto but not including endpoint
    for k in range(19, 23):
        if pd.isnull(siteProps.iloc[1, k]) == False:
            et.SubElement(datasetVar, str(siteProps.iloc[0,k])).text = str(siteProps.iloc[1,k])

    investCount = et.SubElement(datasetVar, "Investigators")
    contactVar = et.SubElement(investCount, "Contact", ID="2")

    processCount = et.SubElement(datasetVar, "Processors")
    contactVar = et.SubElement(processCount, "Contact", ID="1")
    contactVar = et.SubElement(processCount, "Contact", ID="3")

    datasetVar = et.SubElement(datasetsVar, "Dataset")

    et.SubElement(datasetVar, "DatasetType").text = "Water Chemistry"

    for k in range(20, 23):
        if pd.isnull(siteProps.iloc[1, k]) == False:
            # This overwrites that the table assigns a value of IsSSamp = True for water chem since the
            # same columns are re-read to populate the XML for this section
            if str(siteProps.iloc[0,k]) == "IsSSamp":
                et.SubElement(datasetVar, "IsSSamp").text = "False"
            else:
                et.SubElement(datasetVar, str(siteProps.iloc[0,k])).text = str(siteProps.iloc[1,k])


    investCount = et.SubElement(datasetVar, "Investigators")
    contactVar = et.SubElement(investCount, "Contact", ID="2")

    processCount = et.SubElement(datasetVar, "Processors")
    contactVar = et.SubElement(processCount, "Contact", ID="1")
    contactVar = et.SubElement(processCount, "Contact", ID="3")

    # expoort data
    rootWrite = et.ElementTree(root)
    rootWrite.write("batchXML/" + str(siteProps.iloc[1,10]) + "_XML.xml", encoding="UTF-8", xml_declaration=True) # was using column 0 = siteName but now not unique, use 10 = handle
    rootWrite.write("batchTILIA/" + str(siteProps.iloc[1,10]) + "_TILIA.tlx", encoding="UTF-8", xml_declaration=True) # see comment above

# iterate overall database records and generate files (.tlx and .xml)
for i in range(1, len(dfMainTable.index)+1):
    createTilia(i)
    print("Depth_{} is complete" .format(i))
print("Program completed")
